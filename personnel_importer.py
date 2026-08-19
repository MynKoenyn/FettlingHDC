"""
Personnel master-data importer (web upload)
==========================================

Loads personnel from a CSV or Excel file built on the template that
`build_template_csv()` produces (Personnel → Import → Download Template).

This is the web equivalent of the older `import_personnel.py` CLI script.
It reads our own template, so the header row is row 1 and the columns are
known, but headers are still matched loosely (case, spacing and punctuation
are ignored, plus a few synonyms) so a re-saved sheet still loads.

Behaviour, matching the Products importer people already know:

  * Upsert, not append. A row is matched to an existing person by Clock
    Number. Matched rows are updated, the rest inserted, so the same file
    can be re-sent after a correction.
  * A blank cell leaves the existing value alone — it never blanks a field.
  * Department must already exist; it is resolved by name (division-qualified
    when the name repeats across divisions) and never created here, because a
    new department needs a division decision this sheet can't make. The
    division is taken from the resolved department. Unresolved names are
    reported per row and the rest of the row still loads.
  * Pay Group must be one of models.PAY_GROUPS; anything else is reported.
"""

import csv
import io
import re
from datetime import datetime
from decimal import Decimal, InvalidOperation

from app import db
from models import PAY_GROUPS, Department, Division, Personnel


# ── Column headers of the CSV template ───────────────────────────────────────
TEMPLATE_HEADERS = [
    "Clock No",
    "First Name",
    "Surname",
    "Division",
    "Department",
    "Pay Group",
    "Job Grade",
    "Rate",
    "ID Number",
    "Gender",
    "Date Joined",
    "Job Description",
    "Active",
]

# Normalised header → Personnel field (or a resolver key). Exact matches only.
FIELD_ALIASES = {
    "clockno":            "clockno",
    "clocknumber":        "clockno",
    "clock":              "clockno",
    "empno":              "clockno",
    "employeeno":         "clockno",
    "employeenumber":     "clockno",

    "firstname":          "name",
    "firstnames":         "name",
    "name":               "name",
    "names":              "name",

    "surname":            "surname",
    "lastname":           "surname",

    "division":           "division",
    "div":                "division",

    "department":         "department",
    "dept":               "department",

    "paygroup":           "pay_group",
    "payrollgroup":       "pay_group",
    "wagesalary":         "pay_group",
    "wagessalary":        "pay_group",
    "wagesalarygroup":    "pay_group",
    "salarywage":         "pay_group",
    "group":              "pay_group",

    "jobgrade":           "jobgrade",
    "grade":              "jobgrade",
    "hdcgrade":           "jobgrade",

    "rate":               "rate",
    "hourlyrate":         "rate",
    "payrate":            "rate",

    "idnumber":           "id_no",
    "idno":               "id_no",
    "id":                 "id_no",

    "gender":             "gender",
    "sex":                "gender",

    "datejoined":         "joined",
    "joined":             "joined",
    "startdate":          "joined",
    "engagementdate":     "joined",

    "jobdescription":     "job_description",
    "jobdesc":            "job_description",
    "description":        "job_description",
    "position":           "job_description",

    "active":             "status",
    "status":             "status",
    "isactive":           "status",
}

# Columns an export carries that are not imported back — read past them.
IGNORED_HEADERS = {"", "id", "race"}

DATE_FORMATS = [
    "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d", "%d %b %Y", "%d %B %Y",
    "%d.%m.%Y", "%d/%m/%y", "%d-%m-%y",
]

TRUE_WORDS  = {"y", "yes", "true", "t", "1", "active", "x"}
FALSE_WORDS = {"n", "no", "false", "f", "0", "inactive", "terminated", "left"}


# ── Normalisation helpers ────────────────────────────────────────────────────

def norm_header(value):
    return re.sub(r"[^a-z0-9]", "", str(value or "").lower())


def norm_key(value):
    """Case- and space-insensitive key for matching names / codes."""
    return re.sub(r"\s+", " ", str(value or "").strip()).lower()


def clean(value):
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def parse_decimal(value):
    if value is None:
        return None
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value))
    text = re.sub(r"[^0-9,.\-]", "", str(value).strip())
    if not text or text in {"-", ".", ","}:
        return None
    if "," in text and "." in text:
        text = text.replace(",", "")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def parse_date(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    text = clean(value)
    if not text:
        return None
    text = text.split(" ")[0] if re.match(r"^\d", text) and ":" in text else text
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def parse_bool(value):
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    text = re.sub(r"[^a-z0-9]", "", str(value).strip().lower())
    if not text:
        return None
    if text in TRUE_WORDS:
        return True
    if text in FALSE_WORDS:
        return False
    return None


# ── File reading ─────────────────────────────────────────────────────────────

def read_table(file_storage):
    """Read an uploaded CSV/XLSX into (headers, rows)."""
    filename = (file_storage.filename or "").lower()
    data = file_storage.read()

    if filename.endswith((".xlsx", ".xlsm")):
        table = _read_excel(data)
    else:
        table = _read_csv(data)

    start = next(
        (i for i, row in enumerate(table) if any(str(c or "").strip() for c in row)),
        None,
    )
    if start is None:
        return [], []
    return table[start], [r for r in table[start + 1:] if any(str(c or "").strip() for c in r)]


def _read_csv(data):
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = data.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        text = data.decode("utf-8", errors="replace")
    try:
        dialect = csv.Sniffer().sniff(text[:4096], delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    return list(csv.reader(io.StringIO(text), dialect))


def _read_excel(data):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - openpyxl ships in requirements
        raise ValueError("Excel import needs openpyxl installed.") from exc
    workbook = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    return [list(row) for row in workbook.active.iter_rows(values_only=True)]


def map_headers(headers):
    """Returns (index → Personnel field, list of unrecognised headers)."""
    field_map, unknown = {}, []
    for idx, raw in enumerate(headers):
        key = norm_header(raw)
        if key in IGNORED_HEADERS:
            continue
        if key in FIELD_ALIASES:
            field_map[idx] = FIELD_ALIASES[key]
            continue
        unknown.append(str(raw).strip())
    return field_map, unknown


# ── Department / division resolution ─────────────────────────────────────────

def build_division_index():
    """Normalised code or name → Division."""
    index = {}
    for d in Division.query.all():
        if d.code:
            index[norm_key(d.code)] = d
        if d.name:
            index.setdefault(norm_key(d.name), d)
    return index


def build_department_index():
    """
    Returns (by_division, plain, ambiguous):

      by_division : (division_id, normalised dept name) -> Department
                    the reliable lookup once a division is known.
      plain       : normalised dept name -> Department, but only for names
                    that are unique across every division. A name that repeats
                    (e.g. MAINTENANCE under both HDC and HDA) is left out, so a
                    row using it unqualified is flagged rather than guessed.
      ambiguous   : the set of normalised names that repeat across divisions.
    """
    by_division, plain, ambiguous = {}, {}, set()
    for dep in Department.query.all():
        by_division[(dep.division_id, norm_key(dep.name))] = dep
        key = norm_key(dep.name)
        if key in plain and plain[key].division_id != dep.division_id:
            ambiguous.add(key)
        plain.setdefault(key, dep)
    for key in ambiguous:
        plain.pop(key, None)
    return by_division, plain, ambiguous


# ── The import itself ────────────────────────────────────────────────────────

class PersonnelImportResult:
    def __init__(self):
        self.total = 0
        self.created = 0
        self.updated = 0
        self.skipped = 0
        self.unknown_columns = []
        self.row_issues = []      # (row number, name/clock, message)

    @property
    def touched(self):
        return self.created + self.updated


def import_personnel(file_storage):
    """Load a personnel sheet. Returns a PersonnelImportResult.

    Raises ValueError when the file itself is unusable (no header, no Clock
    No column) — a problem with the file rather than with a row.
    """
    result = PersonnelImportResult()

    headers, rows = read_table(file_storage)
    if not headers:
        raise ValueError("That file has no readable header row.")

    field_map, unknown = map_headers(headers)
    result.unknown_columns = unknown

    if "clockno" not in field_map.values():
        raise ValueError(
            "No 'Clock No' column found. Download the template and check the "
            "column headings match."
        )

    divisions = build_division_index()
    depts_by_division, plain_depts, ambiguous_departments = build_department_index()
    pay_groups = {g.lower(): g for g in PAY_GROUPS}

    existing = {norm_key(p.clockno): p for p in Personnel.query.all()}

    for row_no, row in enumerate(rows, start=2):     # row 1 is the header
        result.total += 1
        values = {}
        for idx, field in field_map.items():
            values[field] = row[idx] if idx < len(row) else None

        clockno = clean(values.get("clockno"))
        if not clockno:
            result.skipped += 1
            continue

        person = existing.get(norm_key(clockno))
        is_new = person is None

        # Resolve the department first — a new person can't be created without
        # one (the column is NOT NULL and the division comes from it).
        #
        # Division accepts the code (HDC) or the full name (High Duty Castings).
        # With a division known, the department is looked up within it; without
        # one, only a department name unique across all divisions can resolve.
        dept = None
        div_name = clean(values.get("division"))
        dep_name = clean(values.get("department"))

        division = divisions.get(norm_key(div_name)) if div_name else None
        if div_name and division is None:
            result.row_issues.append((
                row_no, clockno,
                f"Division '{div_name}' does not exist — use a division code (e.g. HDC) or its name."
            ))

        if dep_name and division is not None:
            dept = depts_by_division.get((division.id, norm_key(dep_name)))
            if dept is None:
                result.row_issues.append((
                    row_no, clockno,
                    f"Department '{dep_name}' does not exist in division '{division.code}' — "
                    f"create it first, then re-import."
                ))
        elif dep_name and not div_name:
            # No division named — only a name unique across all divisions can
            # resolve. A name that was specified with an invalid division is
            # never guessed here; that error was already reported above.
            dept = plain_depts.get(norm_key(dep_name))
            if dept is None:
                if norm_key(dep_name) in ambiguous_departments:
                    result.row_issues.append((
                        row_no, clockno,
                        f"Department '{dep_name}' exists in more than one division — "
                        f"fill in the Division column (e.g. HDC) as well."
                    ))
                else:
                    result.row_issues.append((
                        row_no, clockno,
                        f"Department '{dep_name}' does not exist — create it first, then re-import."
                    ))

        if is_new and dept is None:
            result.skipped += 1
            result.row_issues.append((
                row_no, clockno,
                "New personnel need a valid Division/Department — row skipped."
            ))
            continue

        if is_new and not clean(values.get("name")):
            result.skipped += 1
            result.row_issues.append((row_no, clockno, "New personnel need a First Name — row skipped."))
            continue

        if is_new:
            person = Personnel(clockno=clockno)
            db.session.add(person)
            result.created += 1
        else:
            result.updated += 1

        # ---- Text fields. Blank leaves what is already there. --------------
        for field in ("name", "surname", "id_no", "gender", "job_description"):
            text = clean(values.get(field))
            if text:
                setattr(person, field, text)

        grade = clean(values.get("jobgrade"))
        if grade:
            person.jobgrade = grade[:3]

        # ---- Department / division ----------------------------------------
        if dept is not None:
            person.department_id = dept.id
            person.division_id = dept.division_id

        # ---- Pay group -----------------------------------------------------
        pg = clean(values.get("pay_group"))
        if pg:
            matched = pay_groups.get(pg.lower())
            if matched:
                person.pay_group = matched
            else:
                result.row_issues.append((
                    row_no, clockno,
                    f"Pay Group '{pg}' is not one of {', '.join(PAY_GROUPS)} — left unset."
                ))

        # ---- Rate ----------------------------------------------------------
        rate_raw = clean(values.get("rate"))
        if rate_raw:
            rate = parse_decimal(rate_raw)
            if rate is None:
                result.row_issues.append((row_no, clockno, f"Rate '{rate_raw}' is not a number — ignored."))
            elif rate < 0:
                result.row_issues.append((row_no, clockno, "Rate cannot be negative — ignored."))
            else:
                person.rate = rate

        # ---- Date joined ---------------------------------------------------
        joined_raw = clean(values.get("joined"))
        if joined_raw:
            joined = parse_date(joined_raw)
            if joined is None:
                result.row_issues.append((row_no, clockno, f"Date Joined '{joined_raw}' is not a date — ignored."))
            else:
                person.joined = joined

        # ---- Active flag ---------------------------------------------------
        status_raw = clean(values.get("status"))
        flag = parse_bool(status_raw)
        if flag is not None:
            person.status = flag
        elif status_raw:
            result.row_issues.append((row_no, clockno, f"Active '{status_raw}' is not a yes/no value — not applied."))
        elif is_new:
            person.status = True

        # Keep the index current so a file listing the same clock number twice
        # updates the first row's person rather than inserting a duplicate.
        db.session.flush()
        existing.setdefault(norm_key(clockno), person)

    if result.touched:
        db.session.commit()
    else:
        db.session.rollback()

    return result


# ── CSV template / export ────────────────────────────────────────────────────

SAMPLE_ROWS = [
    ["10432", "John", "Smith", "HDC", "FETTLING/WHEELABRATING", "Wages",
     "C2", "48.50", "8001015009087", "Male", "2019-03-01", "Fettler / Grinder", "Yes"],
    ["20515", "Thabo", "Mokoena", "HDA", "UNIT 2", "Salary",
     "D1", "", "9105026009086", "Male", "2021-07-15", "Line Supervisor", "Yes"],
]


def build_template_csv(sample=True):
    """The blank import sheet, with two worked example rows."""
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(TEMPLATE_HEADERS)
    if sample:
        writer.writerows(SAMPLE_ROWS)
    return buffer.getvalue()


def build_export_csv():
    """The current personnel list in template layout — edit it and send it back."""
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(TEMPLATE_HEADERS)
    for p in Personnel.query.order_by(Personnel.name, Personnel.surname).all():
        writer.writerow([
            p.clockno or "",
            p.name or "",
            p.surname or "",
            p.division.code if p.division else "",
            p.department.name if p.department else "",
            p.pay_group or "",
            p.jobgrade or "",
            f"{p.rate:.2f}" if p.rate is not None else "",
            p.id_no or "",
            p.gender or "",
            p.joined.strftime("%Y-%m-%d") if p.joined else "",
            p.job_description or "",
            "No" if p.status is False else "Yes",
        ])
    return buffer.getvalue()
