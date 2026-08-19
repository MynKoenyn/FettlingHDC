"""
Product master-data importer
============================

Loads the product list from a CSV or Excel file built on the template that
`build_template_csv()` produces (Products → Import → Download CSV Template).

Unlike the scrap importer, this reads our own template rather than a
customer's workbook, so the header row is row 1 and the columns are known.
Headers are still matched loosely (case, spacing and punctuation are
ignored, and a few obvious synonyms are accepted) so a sheet re-saved out of
Excel still loads.

Behaviour worth knowing:

  * Upsert, not append. A row is matched to an existing product by Product
    Code, falling back to Name. Matched rows are updated, the rest inserted,
    so the same file can be re-sent after a correction.
  * A blank cell leaves the existing value alone — it never blanks a field.
    That makes a part-filled sheet (say, just Weight) a safe way to top up
    data you already have.
  * Customers and departments must already exist; they are resolved by name
    and never created here, because both need a division that a product
    sheet has no way of telling us. Unresolved names are reported per row
    and the rest of the row still loads.
  * Current Price and Price per kg are deliberately not importable — both
    are worked out for display (from the price list and the weight) and are
    never stored on the product.
"""

import csv
import io
import re
from decimal import Decimal, InvalidOperation

from app import db
from models import PRODUCT_GRADES, Customer, Department, Division, Product


# ── Column headers of the CSV template ───────────────────────────────────────
TEMPLATE_HEADERS = [
    "Name",
    "Product Code",
    "Supplier Code",
    "Simplified Code",
    "Supplier Description",
    "Grade",
    "Drawing Level",
    "Price",
    "Weight (kg)",
    "Primary Customer",
    "Linked Customers",
    "Departments",
    "Barcode",
    "Stock Item",
    "Active",
]

# Normalised header → Product field. Exact matches only.
FIELD_ALIASES = {
    "name":                "name",
    "productname":         "name",
    "description":         "name",

    "productcode":         "product_code",
    "code":                "product_code",
    "castingnumber":       "product_code",
    "castingno":           "product_code",

    "suppliercode":        "supplier_code",
    "customercode":        "supplier_code",
    "machinedpartnumber":  "supplier_code",

    "simplifiedcode":      "simplified_code",
    "simplecode":          "simplified_code",
    "shortcode":           "simplified_code",
    "simplifiedproductcode": "simplified_code",

    "supplierdescription": "supplier_description",
    "suppliordescription": "supplier_description",

    "grade":               "grade",
    "materialgrade":       "grade",

    "drawinglevel":        "drawing_level",
    "drawingrevision":     "drawing_level",
    "drawingrev":          "drawing_level",
    "revision":            "drawing_level",

    "price":               "price",
    "catalogueprice":      "price",
    "unitprice":           "price",

    "weightkg":            "weight",
    "weight":              "weight",
    "castweight":          "weight",
    "kg":                  "weight",
    "mass":                "weight",

    "primarycustomer":     "customer",
    "customer":            "customer",

    "linkedcustomers":     "linked_customers",
    "alsolinkedcustomers": "linked_customers",
    "additionalcustomers": "linked_customers",
    "othercustomers":      "linked_customers",

    "departments":         "departments",
    "department":          "departments",

    "barcode":             "barcode",

    "stockitem":           "is_stock_item",
    "isstockitem":         "is_stock_item",
    "stockeditem":         "is_stock_item",
    "stock":               "is_stock_item",
    "stockstatus":         "is_stock_item",

    "active":              "active",
    "isactive":            "active",
    "inuse":               "active",
}

# Display-only columns. Someone exporting the list and sending it back will
# have these in the sheet; read past them instead of complaining.
IGNORED_HEADERS = {
    "", "id", "currentprice", "pricelistprice", "pricperkg", "priceperkg",
    "rperkg", "pricekg", "total", "totals",
}

# Separators accepted inside a multi-value cell (departments, linked customers)
MULTI_VALUE_SPLIT = re.compile(r"[;|/\n]+")


# ── Normalisation helpers ────────────────────────────────────────────────────

def norm_header(value):
    """'Weight (kg)' → 'weightkg'"""
    return re.sub(r"[^a-z0-9]", "", str(value or "").lower())


def norm_code(value):
    """'1 203 VB2 00' → '1203VB200' — codes ignore spacing and punctuation."""
    return re.sub(r"[^A-Z0-9]", "", str(value or "").upper())


def norm_name(value):
    """Case- and space-insensitive key for matching names."""
    return re.sub(r"\s+", " ", str(value or "").strip()).lower()


def clean(value):
    """A stripped string, or None when the cell is empty."""
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def parse_decimal(value):
    """
    'R 1 234,56' → Decimal('1234.56'). None when the cell is empty or unusable.

    Accepts the comma as a decimal separator (common in local sheets) when it
    is the only separator present; with both, the comma is thousands.
    """
    if value is None:
        return None
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value))

    text = re.sub(r"[^0-9,.\-]", "", str(value).strip())
    if not text or text in {"-", ".", ","}:
        return None

    if "," in text and "." in text:
        text = text.replace(",", "")
    elif "," in text:
        text = text.replace(",", ".")

    try:
        return Decimal(text)
    except InvalidOperation:
        return None


TRUE_WORDS  = {"y", "yes", "true", "t", "1", "active", "stock", "stockitem", "x"}
FALSE_WORDS = {"n", "no", "false", "f", "0", "inactive", "nonstock", "madetoorder"}


def parse_bool(value):
    """
    'Yes' / 'N' / 'TRUE' / 1 → True or False. None when the cell says nothing
    recognisable, which the caller reads as "leave the flag alone".
    """
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)

    text = re.sub(r"[^a-z0-9]", "", str(value).strip().lower())
    if not text:
        return None
    if text in TRUE_WORDS:
        return True
    if text in FALSE_WORDS:
        return False
    return None


def split_multi(value):
    """'Fettling; Machining' → ['Fettling', 'Machining']"""
    text = clean(value)
    if not text:
        return []
    return [part.strip() for part in MULTI_VALUE_SPLIT.split(text) if part.strip()]


# ── File reading ─────────────────────────────────────────────────────────────

def read_table(file_storage):
    """Read an uploaded CSV/XLSX into (headers, rows). Header is the first
    non-empty row — this is our own template, not a customer's workbook."""
    filename = (file_storage.filename or "").lower()
    data = file_storage.read()

    if filename.endswith((".xlsx", ".xlsm")):
        table = _read_excel(data)
    else:
        table = _read_csv(data)

    start = next(
        (i for i, row in enumerate(table) if any(str(c or "").strip() for c in row)),
        None,
    )
    if start is None:
        return [], []

    headers = table[start]
    rows = [r for r in table[start + 1:] if any(str(c or "").strip() for c in r)]
    return headers, rows


def _read_csv(data):
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = data.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        text = data.decode("utf-8", errors="replace")

    try:
        dialect = csv.Sniffer().sniff(text[:4096], delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel

    return list(csv.reader(io.StringIO(text), dialect))


def _read_excel(data):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - openpyxl ships in requirements
        raise ValueError("Excel import needs openpyxl installed.") from exc

    workbook = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    return [list(row) for row in workbook.active.iter_rows(values_only=True)]


def map_headers(headers):
    """Returns (index → Product field, list of unrecognised headers)."""
    field_map, unknown = {}, []
    for idx, raw in enumerate(headers):
        key = norm_header(raw)
        if key in IGNORED_HEADERS:
            continue
        if key in FIELD_ALIASES:
            field_map[idx] = FIELD_ALIASES[key]
            continue
        unknown.append(str(raw).strip())
    return field_map, unknown


# ── Lookups for the columns that reference other records ─────────────────────

def build_customer_index():
    """Normalised name → Customer."""
    return {norm_name(c.name): c for c in Customer.query.all() if c.name}


def build_department_index():
    """
    Keys for a department: its plain name, and the division-qualified forms
    the template writes ('HDC — Fettling', 'HDC/Fettling').

    Department names repeat across divisions, so a plain name that is not
    unique is left out — those rows have to say which division they mean.
    """
    index = {}
    ambiguous = set()

    for dep in Department.query.join(Division, Department.division_id == Division.id).all():
        plain = norm_name(dep.name)
        if plain in index:
            ambiguous.add(plain)
        index.setdefault(plain, dep)

        code = dep.division.code if dep.division else None
        if code:
            for qualified in (f"{code} — {dep.name}", f"{code} - {dep.name}",
                              f"{code}/{dep.name}", f"{code} {dep.name}"):
                index[norm_name(qualified)] = dep

    for key in ambiguous:
        index.pop(key, None)
    return index, ambiguous


# ── The import itself ────────────────────────────────────────────────────────

class ProductImportResult:
    """What one upload did, for the flash message and the import screen."""

    def __init__(self):
        self.total = 0
        self.created = 0
        self.updated = 0
        self.skipped = 0
        self.unknown_columns = []
        self.row_issues = []      # (row number, product name, message)

    @property
    def touched(self):
        return self.created + self.updated


def import_products(file_storage):
    """
    Load a product sheet. Returns a ProductImportResult.

    Raises ValueError when the file itself is unusable (no header, no Name
    column) — a problem with the file rather than with a row.
    """
    result = ProductImportResult()

    headers, rows = read_table(file_storage)
    if not headers:
        raise ValueError("That file has no readable header row.")

    field_map, unknown = map_headers(headers)
    result.unknown_columns = unknown

    # A row has to be identifiable. Name alone is enough, and so is Product
    # Code alone — a sheet of codes plus one column is a normal way to top up
    # a single field across the catalogue.
    if not {"name", "product_code"} & set(field_map.values()):
        raise ValueError(
            "No 'Name' or 'Product Code' column found. Download the CSV "
            "template and check the column headings match."
        )

    customers = build_customer_index()
    departments, ambiguous_departments = build_department_index()
    grades = {g.upper(): g for g in PRODUCT_GRADES}

    # Existing products, indexed the same two ways a row is matched.
    by_code, by_name = {}, {}
    for product in Product.query.all():
        code = norm_code(product.product_code)
        if code:
            by_code.setdefault(code, product)
        by_name.setdefault(norm_name(product.name), product)

    for row_no, row in enumerate(rows, start=2):     # row 1 is the header
        result.total += 1

        values = {}
        for idx, field in field_map.items():
            values[field] = row[idx] if idx < len(row) else None

        name = clean(values.get("name"))
        code = clean(values.get("product_code"))

        if not name and not code:
            result.skipped += 1
            continue

        product = by_code.get(norm_code(code)) if code else None
        if product is None and name:
            product = by_name.get(norm_name(name))

        if product is None:
            if not name:
                result.skipped += 1
                result.row_issues.append(
                    (row_no, code, "New product needs a Name — row skipped.")
                )
                continue
            product = Product(name=name)
            db.session.add(product)
            result.created += 1
            is_new = True
        else:
            result.updated += 1
            is_new = False

        # ---- Plain text fields. Blank leaves what is already there. --------
        if name:
            product.name = name
        for field in ("product_code", "supplier_code", "simplified_code",
                      "supplier_description", "barcode", "drawing_level"):
            text = clean(values.get(field))
            if text:
                setattr(product, field, text)

        # ---- Grade — must be one we cast in --------------------------------
        grade = clean(values.get("grade"))
        if grade:
            matched = grades.get(grade.upper())
            if matched:
                product.grade = matched
            else:
                result.row_issues.append((
                    row_no, name,
                    f"Grade '{grade}' is not one of {', '.join(PRODUCT_GRADES)} — left unset."
                ))

        # ---- Numbers -------------------------------------------------------
        for field, label in (("price", "Price"), ("weight", "Weight")):
            raw = clean(values.get(field))
            if not raw:
                continue
            number = parse_decimal(raw)
            if number is None:
                result.row_issues.append((row_no, name, f"{label} '{raw}' is not a number — ignored."))
            elif number < 0:
                result.row_issues.append((row_no, name, f"{label} cannot be negative — ignored."))
            else:
                setattr(product, field, number)

        # ---- Flags. A blank cell leaves the flag as it is; a new product
        #      defaults to a stock item that is active. ------------------------
        for field, label in (("is_stock_item", "Stock Item"), ("active", "Active")):
            raw = clean(values.get(field))
            flag = parse_bool(raw)
            if flag is not None:
                setattr(product, field, flag)
            elif raw:
                result.row_issues.append((
                    row_no, name,
                    f"{label} '{raw}' is not a yes/no value — not applied."
                ))
            elif is_new:
                setattr(product, field, True)

        # ---- Primary customer ----------------------------------------------
        primary = clean(values.get("customer"))
        if primary:
            match = customers.get(norm_name(primary))
            if match:
                product.customer_id = match.id
            else:
                result.row_issues.append((
                    row_no, name,
                    f"Customer '{primary}' does not exist — add it first, then re-import."
                ))

        # ---- Linked customers (replaces the set when the cell has a value) --
        linked_names = split_multi(values.get("linked_customers"))
        if linked_names:
            linked, missing = [], []
            for entry in linked_names:
                match = customers.get(norm_name(entry))
                (linked if match else missing).append(match or entry)
            product.linked_customers = linked
            if missing:
                result.row_issues.append((
                    row_no, name,
                    "Linked customer(s) not found: " + ", ".join(missing)
                ))

        # ---- Departments (replaces the set when the cell has a value) -------
        department_names = split_multi(values.get("departments"))
        if department_names:
            resolved, missing = [], []
            for entry in department_names:
                match = departments.get(norm_name(entry))
                if match:
                    resolved.append(match)
                elif norm_name(entry) in ambiguous_departments:
                    missing.append(f"{entry} (in more than one division — write it as 'HDC — {entry}')")
                else:
                    missing.append(entry)
            product.departments = resolved
            if missing:
                result.row_issues.append((
                    row_no, name, "Department(s) not found: " + ", ".join(missing)
                ))

        # Keep the indexes current so a file listing the same product twice
        # updates the first row's product rather than inserting a second.
        db.session.flush()
        if is_new:
            by_name.setdefault(norm_name(product.name), product)
        if product.product_code:
            by_code.setdefault(norm_code(product.product_code), product)

    if result.touched:
        db.session.commit()
    else:
        db.session.rollback()

    return result


# ── CSV template ─────────────────────────────────────────────────────────────

SAMPLE_ROWS = [
    ["Brake Caliper Housing", "BCH-001", "SUP-4471", "CH-001", "Caliper housing LH",
     "SG50", "AA", "1250.00", "12.500", "ACME Motors", "", "Fettling; Machining", "",
     "Yes", "Yes"],
    ["Flywheel Cover", "FWC-220", "SUP-9903", "FC-220", "Cover, flywheel",
     "GG25", "OOO", "480.50", "6.250", "ACME Motors", "Beta Foundry Supplies",
     "HDC — Fettling", "6001234567890", "No", "Yes"],
]


def build_template_csv(sample=True):
    """The blank import sheet, with two worked example rows."""
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(TEMPLATE_HEADERS)

    if sample:
        writer.writerows(SAMPLE_ROWS)

    return buffer.getvalue()


def build_export_csv():
    """The current product list in template layout — edit it and send it back."""
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(TEMPLATE_HEADERS)

    for p in Product.query.order_by(Product.name).all():
        writer.writerow([
            p.name,
            p.product_code or "",
            p.supplier_code or "",
            p.simplified_code or "",
            p.supplier_description or "",
            p.grade or "",
            p.drawing_level or "",
            f"{p.price:.2f}" if p.price is not None else "",
            f"{p.weight:.3f}" if p.weight is not None else "",
            p.customer.name if p.customer else "",
            "; ".join(c.name for c in p.linked_customers),
            "; ".join(
                f"{d.division.code} — {d.name}" if d.division else d.name
                for d in p.departments
            ),
            p.barcode or "",
            "No" if p.is_stock_item is False else "Yes",
            "No" if p.active is False else "Yes",
        ])

    return buffer.getvalue()
