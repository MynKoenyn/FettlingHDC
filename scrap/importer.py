"""
Scrap module — external report importer
=======================================

Reads a customer's reject report (CSV or Excel) and turns each row into a
ScrapEntry plus its defect breakdown.

The customer's sheet is not written for us: headers are multi-line, part
numbers carry spaces, dates come through as "09-Jan-26", and derived columns
(% Reject, Monthly Summary) sit alongside the real ones. So rather than
demanding an exact layout, the importer normalises every header and matches
it against the known core fields and the defect catalogue — the file can be
saved straight out of the customer's workbook and still load.

Re-importing the same report is safe: every external row gets a dedupe key
of customer + part + casting + batch + date, and a row whose key is already
taken is never loaded blind. It is parked as a ScrapPendingDuplicate for
someone to allow or reject — the same part and batch can legitimately be
rejected twice, so that call belongs to a person, not to the importer.
"""

import csv
import hashlib
import io
import json
import re
from datetime import date, datetime

from app import db
from models import Customer, Product, product_customers
from scrap.models import (
    CLASH_EXISTING,
    CLASH_FILE,
    DUP_ALLOWED,
    DUP_PENDING,
    DUP_REJECTED,
    SCOPE_EXTERNAL,
    SOURCE_EXTERNAL,
    ScrapDefect,
    ScrapEntry,
    ScrapEntryDefect,
    ScrapImportBatch,
    ScrapPendingDuplicate,
    active_defects,
)


# ── Column headers of the CSV template ───────────────────────────────────────
CORE_TEMPLATE_HEADERS = [
    "Machined Part Number",
    "Casting Number",
    "Batch #",
    "Reject Date",
    "QTY Booked",
    "Receiving QTY",
    "QTY Scrap",
    "QTY Machined",
]

# Normalised header → ScrapEntry field. Exact matches only.
CORE_FIELD_ALIASES = {
    "machinedpartnumber": "machined_part_no",
    "machinedpartno":     "machined_part_no",
    "machinedpart":       "machined_part_no",
    "machinedpn":         "machined_part_no",
    "partnumber":         "machined_part_no",
    "partno":             "machined_part_no",

    "castingnumber":      "casting_no",
    "castingno":          "casting_no",
    "casting":            "casting_no",
    "castingpartnumber":  "casting_no",

    "batch":              "batch_no",
    "batchno":            "batch_no",
    "batchnumber":        "batch_no",
    "batchcode":          "batch_no",

    "rejectdate":         "entry_date",
    "date":               "entry_date",
    "scrapdate":          "entry_date",
    "dateofreject":       "entry_date",

    "qtybooked":          "qty_booked",
    "quantitybooked":     "qty_booked",
    "booked":             "qty_booked",

    "receivingqty":       "qty_received",
    "receivedqty":        "qty_received",
    "qtyreceived":        "qty_received",
    "receivingquantity":  "qty_received",

    "qtyscrap":           "qty_scrap",
    "scrapqty":           "qty_scrap",
    "qtyscrapped":        "qty_scrap",
    "quantityscrap":      "qty_scrap",
    "scrap":              "qty_scrap",
    "rejectqty":          "qty_scrap",
    "qtyreject":          "qty_scrap",

    "qtymachined":        "qty_machined",
    "machinedqty":        "qty_machined",
    "quantitymachined":   "qty_machined",
    "qtyproduced":        "qty_machined",
    "producedqty":        "qty_machined",

    "notes":              "notes",
    "note":               "notes",
    "comment":            "notes",
    "comments":           "notes",
    "remarks":            "notes",
}

# Derived / summary columns on the customer's sheet — read past them.
IGNORED_HEADERS = {
    "", "reject", "rejectpercent", "rejectpercentage", "percentreject",
    "monthlysummary", "yearlysummary", "summary", "total", "totals",
    "grandtotal", "runningtotal", "checksum", "difference",
}

DATE_FORMATS = [
    "%d-%b-%y", "%d-%b-%Y", "%d %b %y", "%d %b %Y",
    "%d-%B-%y", "%d-%B-%Y", "%d %B %Y",
    "%Y-%m-%d", "%Y/%m/%d",
    "%d/%m/%Y", "%d/%m/%y", "%d-%m-%Y", "%d-%m-%y",
    "%d.%m.%Y", "%d.%m.%y",
]


# ── Normalisation helpers ────────────────────────────────────────────────────

def norm_header(value):
    """'Non Clean-up (NC)\\nRaw casting…' → 'noncleanupncrawcasting…'"""
    return re.sub(r"[^a-z0-9]", "", str(value or "").lower())


def norm_code(value):
    """'1 203 VB2 00' → '1203VB200' — part numbers ignore spacing/punctuation."""
    return re.sub(r"[^A-Z0-9]", "", str(value or "").upper())


def parse_int(value):
    """Blank cells, dashes and stray text all read as 0."""
    if value is None:
        return 0
    if isinstance(value, bool):
        return 0
    if isinstance(value, (int, float)):
        return int(round(value))
    text = str(value).strip().replace(",", "").replace(" ", "")
    if not text or text in {"-", "–", "—", "n/a", "N/A"}:
        return 0
    try:
        return int(round(float(text)))
    except ValueError:
        return 0


# Month spellings the customer's sheet uses that Python's %b/%B won't match —
# "Sept" is neither the 3-letter abbreviation nor the full month name.
MONTH_ALIASES = {
    "sept": "sep",
}


def _normalise_month(text):
    def replace(match):
        word = match.group(0)
        return MONTH_ALIASES.get(word.lower(), word)
    return re.sub(r"[A-Za-z]+", replace, text)


def parse_date(value):
    """Accept Excel date cells and the customer's text formats. None if unusable."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    for candidate in (text, _normalise_month(text)):
        for fmt in DATE_FORMATS:
            try:
                return datetime.strptime(candidate, fmt).date()
            except ValueError:
                continue
    return None


def make_dedupe_key(customer_id, machined_part_no, casting_no, batch_no, entry_date):
    """Stable natural key for an external row."""
    raw = "|".join([
        str(customer_id or ""),
        norm_code(machined_part_no),
        norm_code(casting_no),
        norm_code(batch_no),
        entry_date.isoformat() if entry_date else "",
    ])
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()


def unique_dedupe_key(base_key):
    """
    A free key for a duplicate someone chose to keep.

    dedupe_key is unique, so an allowed repeat cannot reuse the original. It
    is stored under a numbered variant instead — the row still lands, and
    re-sending the same file later still recognises the original as taken.
    """
    if not ScrapEntry.query.filter_by(dedupe_key=base_key).first():
        return base_key

    occurrence = 2
    while True:
        candidate = hashlib.sha1(f"{base_key}|dup{occurrence}".encode("utf-8")).hexdigest()
        if not ScrapEntry.query.filter_by(dedupe_key=candidate).first():
            return candidate
        occurrence += 1


# ── Header → field mapping ───────────────────────────────────────────────────

def build_defect_lookup(defects=None):
    """
    Normalised header text → ScrapDefect, plus the prefix candidates used when
    a customer's header carries the defect name *and* its description.
    """
    defects = defects if defects is not None else active_defects(SCOPE_EXTERNAL)

    exact = {}
    prefixes = []   # (normalised name, defect) — longest match wins
    for d in defects:
        keys = {norm_header(d.code), norm_header(d.name)}
        if d.description:
            keys.add(norm_header(f"{d.name} {d.description}"))
            keys.add(norm_header(d.description))
        for alias in d.alias_list:
            keys.add(norm_header(alias))
        for key in keys:
            if key:
                exact.setdefault(key, d)

        name_key = norm_header(d.name)
        if name_key:
            prefixes.append((name_key, d))

    prefixes.sort(key=lambda pair: len(pair[0]), reverse=True)
    return exact, prefixes


def map_headers(headers, defects=None):
    """
    Work out what each column of the file is.

    Returns (core_map, defect_map, unknown):
      core_map   index → ScrapEntry field name
      defect_map index → ScrapDefect
      unknown    original headers that were not recognised (ignored on import)
    """
    exact_defects, defect_prefixes = build_defect_lookup(defects)

    core_map, defect_map, unknown = {}, {}, []

    for idx, raw in enumerate(headers):
        key = norm_header(raw)

        if key in IGNORED_HEADERS:
            continue
        if key in CORE_FIELD_ALIASES:
            core_map[idx] = CORE_FIELD_ALIASES[key]
            continue
        if key in exact_defects:
            defect_map[idx] = exact_defects[key]
            continue

        # Customer headers often read "Face (BF) Blow hole on friction face" —
        # match on the defect name the header starts with.
        matched = next((d for prefix, d in defect_prefixes if key.startswith(prefix)), None)
        if matched:
            defect_map[idx] = matched
            continue

        unknown.append(str(raw).strip())

    return core_map, defect_map, unknown


# ── File readers ─────────────────────────────────────────────────────────────

def read_rows(file_storage):
    """
    Read an uploaded CSV/XLSX into (headers, rows).

    Leading blank/title rows are skipped — the header row is the first one
    that looks like the report's own header.
    """
    filename = (file_storage.filename or "").lower()
    data = file_storage.read()

    if filename.endswith((".xlsx", ".xlsm")):
        table = _read_excel(data)
    else:
        table = _read_csv(data)

    return _split_header(table)


def _read_csv(data):
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = data.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        text = data.decode("utf-8", errors="replace")

    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel

    return [row for row in csv.reader(io.StringIO(text), dialect)]


def _read_excel(data):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - openpyxl ships in requirements
        raise ValueError("Excel import needs openpyxl installed.") from exc

    workbook = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    sheet = workbook.active
    return [list(row) for row in sheet.iter_rows(values_only=True)]


def _split_header(table):
    """Find the header row and return (headers, data_rows)."""
    header_idx = None
    for idx, row in enumerate(table[:25]):
        key_text = " ".join(norm_header(c) for c in row)
        hits = sum(1 for alias in ("machinedpart", "casting", "batch", "rejectdate", "qtyscrap")
                   if alias in key_text)
        if hits >= 2:
            header_idx = idx
            break

    if header_idx is None:
        # No recognisable header — fall back to the first non-empty row.
        header_idx = next(
            (i for i, row in enumerate(table) if any(str(c or "").strip() for c in row)),
            0,
        )

    headers = table[header_idx] if table else []
    rows = [r for r in table[header_idx + 1:] if any(str(c or "").strip() for c in r)]
    return headers, rows


# ── Product matching ─────────────────────────────────────────────────────────

def build_product_index(customer_id=None):
    """
    Normalised part code → Product.

    A product is indexed under its product code, supplier code, simplified
    code, supplier description, name and barcode. Supplier description is
    included because that is where the customer's full casting-number format
    (e.g. "1 202 443 01") lives on this catalogue — product/simplified code
    only holds the short form ("443 01"), which never appears verbatim on a
    customer's reject report. Where two products share a code, the one
    belonging to the customer being imported wins.
    """
    linked_ids = set()
    if customer_id:
        linked_ids = {
            row.product_id for row in
            db.session.query(product_customers.c.product_id)
            .filter(product_customers.c.customer_id == customer_id)
            .all()
        }

    def belongs(product):
        return bool(customer_id) and (
            product.customer_id == customer_id or product.id in linked_ids
        )

    index = {}
    for product in Product.query.all():
        for raw in (product.product_code, product.supplier_code,
                    product.simplified_code, product.supplier_description,
                    product.name, product.barcode):
            key = norm_code(raw)
            if not key:
                continue
            current = index.get(key)
            if current is None or (belongs(product) and not belongs(current)):
                index[key] = product
    return index


def match_product(index, casting_no, machined_part_no):
    """Casting number first — it is what HDC supplied — then the machined part."""
    for candidate in (casting_no, machined_part_no):
        product = index.get(norm_code(candidate))
        if product:
            return product
    return None


# ── The import itself ────────────────────────────────────────────────────────

class ImportResult:
    """What one upload did, for the flash message and the batch screen."""

    def __init__(self):
        self.total = 0
        self.imported = 0
        self.duplicates = 0
        self.skipped = 0
        self.unmatched = 0
        self.unknown_columns = []
        self.warnings = []
        self.batch = None

    @property
    def notes_text(self):
        parts = []
        if self.unknown_columns:
            parts.append("Unrecognised columns ignored: " + ", ".join(self.unknown_columns))
        parts.extend(self.warnings)
        return "\n".join(parts) or None


def import_external_report(file_storage, customer_id, user_id=None, default_date=None):
    """
    Load a customer scrap report.

    Rows whose key is already taken (same customer + part + casting + batch +
    date) are held back as pending duplicates rather than dropped, so someone
    can allow or reject each one. Rows whose part number matches no Product
    still import — they are flagged unmatched so the product link can be fixed
    later without re-uploading.
    """
    result = ImportResult()

    headers, rows = read_rows(file_storage)
    if not headers:
        raise ValueError("That file has no readable header row.")

    defects = active_defects(SCOPE_EXTERNAL)
    core_map, defect_map, unknown = map_headers(headers, defects)
    result.unknown_columns = unknown

    if "qty_scrap" not in core_map.values() and not defect_map:
        raise ValueError(
            "No scrap quantity columns found. Download the CSV template and "
            "check the column headings match."
        )

    customer = Customer.query.get(customer_id) if customer_id else None
    product_index = build_product_index(customer_id)

    batch = ScrapImportBatch(
        filename=file_storage.filename,
        customer_id=customer_id,
        imported_by=user_id,
    )
    db.session.add(batch)
    db.session.flush()          # need batch.id for the entries
    result.batch = batch

    # Keys added earlier in this same file — catches duplicates within one upload
    seen_keys = set()

    for row_no, row in enumerate(rows, start=1):
        result.total += 1
        values = {}

        for idx, field in core_map.items():
            raw = row[idx] if idx < len(row) else None
            if field == "entry_date":
                values[field] = parse_date(raw)
            elif field in ("machined_part_no", "casting_no", "batch_no", "notes"):
                values[field] = (str(raw).strip() if raw not in (None, "") else None)
            else:
                values[field] = parse_int(raw)

        defect_qtys = {}
        for idx, defect in defect_map.items():
            qty = parse_int(row[idx] if idx < len(row) else None)
            if qty:
                defect_qtys[defect.id] = defect_qtys.get(defect.id, 0) + qty

        entry_date = values.get("entry_date") or default_date
        qty_scrap = values.get("qty_scrap") or 0
        defect_total = sum(defect_qtys.values())

        # A sheet may only fill the defect columns — take their sum as the scrap qty.
        if not qty_scrap and defect_total:
            qty_scrap = defect_total

        if not entry_date:
            result.skipped += 1
            result.warnings.append(f"Row {row_no}: no usable reject date — skipped.")
            continue
        if not qty_scrap:
            result.skipped += 1
            continue

        qty_machined = values.get("qty_machined") or 0
        if not qty_machined:
            # The sheet's QTY Machined is booked + scrap; derive it when absent.
            qty_machined = (values.get("qty_booked") or 0) + qty_scrap

        dedupe_key = make_dedupe_key(
            customer_id,
            values.get("machined_part_no"),
            values.get("casting_no"),
            values.get("batch_no"),
            entry_date,
        )

        # Already loaded — park the row for a decision rather than dropping it.
        # Earlier rows of this same file are flushed, so one query catches both
        # a repeat inside the file and one imported previously.
        clash_entry = ScrapEntry.query.filter_by(dedupe_key=dedupe_key).first()
        if dedupe_key in seen_keys or clash_entry:
            result.duplicates += 1
            db.session.add(ScrapPendingDuplicate(
                batch_id=batch.id,
                source_row=row_no,
                dedupe_key=dedupe_key,
                clash=CLASH_FILE if dedupe_key in seen_keys else CLASH_EXISTING,
                existing_entry_id=clash_entry.id if clash_entry else None,
                customer_id=customer_id,
                entry_date=entry_date,
                machined_part_no=values.get("machined_part_no"),
                casting_no=values.get("casting_no"),
                batch_no=values.get("batch_no"),
                qty_booked=values.get("qty_booked") or 0,
                qty_received=values.get("qty_received") or 0,
                qty_scrap=qty_scrap,
                qty_machined=qty_machined,
                notes=values.get("notes"),
                defects_json=json.dumps({str(k): v for k, v in defect_qtys.items()}),
                status=DUP_PENDING,
            ))
            continue
        seen_keys.add(dedupe_key)

        product = match_product(
            product_index, values.get("casting_no"), values.get("machined_part_no")
        )
        if product is None:
            result.unmatched += 1

        entry = ScrapEntry(
            source=SOURCE_EXTERNAL,
            entry_date=entry_date,
            customer_id=customer_id,
            product_id=product.id if product else None,
            machined_part_no=values.get("machined_part_no"),
            casting_no=values.get("casting_no"),
            batch_no=values.get("batch_no"),
            qty_booked=values.get("qty_booked") or 0,
            qty_received=values.get("qty_received") or 0,
            qty_scrap=qty_scrap,
            qty_machined=qty_machined,
            notes=values.get("notes"),
            batch_id=batch.id,
            source_row=row_no,
            dedupe_key=dedupe_key,
            created_by=user_id,
        )
        db.session.add(entry)
        db.session.flush()

        for defect_id, qty in defect_qtys.items():
            db.session.add(ScrapEntryDefect(entry_id=entry.id, defect_id=defect_id, qty=qty))

        if defect_total and defect_total != qty_scrap:
            result.warnings.append(
                f"Row {row_no}: defect breakdown ({defect_total}) does not match "
                f"QTY Scrap ({qty_scrap})."
            )

        result.imported += 1

    batch.rows_total     = result.total
    batch.rows_imported  = result.imported
    batch.rows_duplicate = result.duplicates
    batch.rows_skipped   = result.skipped
    batch.rows_unmatched = result.unmatched
    batch.notes          = result.notes_text

    if result.imported == 0 and result.duplicates == 0:
        # Nothing landed — don't leave an empty batch behind.
        db.session.rollback()
        result.batch = None
    else:
        db.session.commit()

    if customer and result.imported:
        result.warnings.append(f"Imported against customer {customer.name}.")

    return result


# ── Resolving a held duplicate ───────────────────────────────────────────────

def allow_duplicate(pending, user_id=None):
    """
    Import a held row after all. Returns the new ScrapEntry, or None if the
    row was already decided.

    Product matching runs again here rather than reusing whatever the import
    found — a part missing then may well have been loaded since.

    The caller commits.
    """
    if pending.status != DUP_PENDING:
        return None

    index = build_product_index(pending.customer_id)
    product = match_product(index, pending.casting_no, pending.machined_part_no)

    entry = ScrapEntry(
        source=SOURCE_EXTERNAL,
        entry_date=pending.entry_date,
        customer_id=pending.customer_id,
        product_id=product.id if product else None,
        machined_part_no=pending.machined_part_no,
        casting_no=pending.casting_no,
        batch_no=pending.batch_no,
        qty_booked=pending.qty_booked or 0,
        qty_received=pending.qty_received or 0,
        qty_scrap=pending.qty_scrap or 0,
        qty_machined=pending.qty_machined or 0,
        notes=pending.notes,
        batch_id=pending.batch_id,
        source_row=pending.source_row,
        dedupe_key=unique_dedupe_key(pending.dedupe_key),
        created_by=user_id,
    )
    db.session.add(entry)
    db.session.flush()

    for defect_id, qty in pending.defect_qtys.items():
        db.session.add(ScrapEntryDefect(entry_id=entry.id, defect_id=defect_id, qty=qty))

    pending.status           = DUP_ALLOWED
    pending.resolved_at      = datetime.now()
    pending.resolved_by      = user_id
    pending.created_entry_id = entry.id

    # The batch counters describe what it actually put in the system
    batch = pending.batch
    if batch:
        batch.rows_imported = (batch.rows_imported or 0) + 1
        if product is None:
            batch.rows_unmatched = (batch.rows_unmatched or 0) + 1

    return entry


def reject_duplicate(pending, user_id=None):
    """
    Discard a held row. Who decided, and when, stays on the batch as the
    record of why the file's row count and the entry count differ.

    The caller commits.
    """
    if pending.status != DUP_PENDING:
        return False

    pending.status      = DUP_REJECTED
    pending.resolved_at = datetime.now()
    pending.resolved_by = user_id
    return True


# ── CSV template ─────────────────────────────────────────────────────────────

def template_headers(defects=None):
    defects = defects if defects is not None else active_defects(SCOPE_EXTERNAL)
    return CORE_TEMPLATE_HEADERS + [d.name for d in defects]


def build_template_csv(defects=None, sample=True):
    """The blank import template, with one worked example row."""
    defects = defects if defects is not None else active_defects(SCOPE_EXTERNAL)
    headers = template_headers(defects)

    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(headers)

    if sample:
        # Mirrors the first line of the customer report: 320 booked + 18 scrap
        # = 338 machined, the 18 split across two defect columns.
        example = {"OB": 4, "NC": 4, "BL": 8, "IF": 1, "MET": 1}
        row = ["1 103 VB2 01", "1 203 VB2 00", "371I", "09-Jan-26", 320, "", 18, 338]
        row += [example.get(d.code, "") for d in defects]
        writer.writerow(row)

    return buffer.getvalue()
